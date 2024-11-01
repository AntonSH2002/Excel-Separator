import openpyxl
from copy import copy


def save_chunk_with_style(sheet, output_file, start_row, end_row):
    # Создаем новую книгу в режиме `write_only` для записи
    new_wb = openpyxl.Workbook(write_only=True)
    new_sheet = new_wb.create_sheet("Sheet1")

    # Копируем нужные строки из основного листа
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row):
        new_row = []
        for cell in row:
            new_cell = openpyxl.cell.WriteOnlyCell(new_sheet, value=cell.value)
            # Копируем стили
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = cell.number_format
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)
            new_row.append(new_cell)

        new_sheet.append(new_row)

    # Сохраняем фрагмент в новый файл
    new_wb.save(output_file)


def split_excel_with_style(file_path, output_prefix, rows_per_file):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    total_rows = sheet.max_row
    num_splits = (total_rows // rows_per_file) + (1 if total_rows % rows_per_file > 0 else 0)

    for i in range(num_splits):
        start_row = i * rows_per_file + 1
        end_row = min(start_row + rows_per_file - 1, total_rows)
        output_file = f"{output_prefix}_part_{i + 1}.xlsx"

        # Сохраняем каждый фрагмент
        save_chunk_with_style(sheet, output_file, start_row, end_row)
        print(f"Сохранен файл: {output_file}")

    wb.close()


# Пример использования
file_path = "table.xlsx"
output_prefix = "tab"
rows_per_file = 50000  # Количество строк на один файл

split_excel_with_style(file_path, output_prefix, rows_per_file)
